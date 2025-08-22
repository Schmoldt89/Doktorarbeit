# für HVR1 und HVR2 die Ruhe Mittelwerte für M (13) rausfiltern

from asyncio.windows_events import NULL
import os
from os import listdir
from genericpath import isfile
from ntpath import join
from openpyxl import load_workbook
from openpyxl import Workbook
import math
import numpy as np
from openpyxl.styles import PatternFill
import statistics
from scipy.stats import t
#--------------------------------
#-----global vars----------------
#--------------------------------

hvr1 = "HVR 1"
hvr2 = "HVR 2"
hvr3 = "HVR 3"
lufu = "Lufu"
hvcr = "HCVR"
datapath = os.path.dirname(os.path.realpath(__file__))
globalcount=0

#--------------------------------
#-------functions----------------
#--------------------------------

def make_new_folder(foldername):
    path = os.path.dirname(os.path.realpath(__file__)) + "\\" + foldername
    if not os.path.exists(path):
        os.mkdir(path)

def get_filenames(path):
    return [f for f in listdir(path) if isfile(join(path, f))]

#Funktion, die Zieldateien erstellt und das manipulierte Dataset speichert
def make_result_dirs(path, data_listofworkbooks):
    make_new_folder("Result")
    filenames = get_filenames(path+"\Input")
    i =0
    for filename in filenames:
        if not os.path.exists(path + "\\Result" + "\\" + filename[:-5] + "_ausgewertet" + filename[-5:]):
            data_listofworkbooks[i].save(path + "\\Result" + "\\" + filename[:-5] + "_ausgewertet" + filename[-5:])
            i +=1

def get_spreadsheets(path):
    onlyfiles = get_filenames(path)
    result=[]
    for file in onlyfiles:
        temp =load_workbook(path+'\\'+file)
        result.append(temp)
    return result

def color_field(datasheet, color, limiter_row_beg, limiter_col_beg, limiter_row_end, limiter_col_end):
    for row in datasheet.iter_rows(min_row=limiter_row_beg, min_col=limiter_col_beg, max_row=limiter_row_end, max_col=limiter_col_end):
        for col in row:
            col.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
def truedata_eval_hvr1(datasheet):
    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    if len(limiters) != 4:
        print(globalcount)
        raise Exception("incorrect number of limiters HVR1")
    limiter_beg1 = limiters[0]
    limiter_beg2 = limiters[2]
    limiter_end1 = limiters[1]
    limiter_end2 = limiters[3]

    k2_arr = []
    p2_arr = []
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=11, max_col=11):
        k2_arr.append(val[0].value)
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=16, max_col=16):
        p2_arr.append(val[0].value)

    datasheet['A6'] = "MitRuheVE" # Average(limiter_beg1, limiter_end1), Column K
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=11, max_col=11):
        sumvar += val[0].value
        iterate += 1
    datasheet['B6'] = sumvar/iterate

    datasheet['A7'] = "MitRuheBf" # Average(limiter_beg1, limiter_end1), Column L
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=12, max_col=12):
        sumvar += val[0].value
        iterate += 1
    datasheet['B7'] = sumvar/iterate

    datasheet['A8'] = "MitRuheVtL" # Average(limiter_beg1, limiter_end1), Column L
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=13, max_col=13):
        sumvar += val[0].value
        iterate += 1
    datasheet['B8'] = sumvar/iterate

    datasheet['A9'] = "MitRuheSpO2" # Average(limiter_beg1, limiter_end1), Column P
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=16, max_col=16):
        sumvar += val[0].value
        iterate += 1
    datasheet['B9'] = sumvar/iterate

    datasheet['A10'] = "MitRuhePetO2" # Average(limiter_beg1, limiter_end1), Column Q
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=17, max_col=17):
        sumvar += val[0].value
        iterate += 1
    datasheet['B10']= sumvar/iterate

    datasheet['A11'] = "MitRuhePetCO2" # Average(limiter_beg1, limiter_end1), Column R
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg1, max_row=limiter_end1, min_col=18, max_col=18):
        sumvar += val[0].value
        iterate += 1
    datasheet['B11']= sumvar/iterate
    
    datasheet['A12'] = "MitHVRPetCO2" # Average(limiter_beg2, limiter_end2), Column R
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        sumvar += val[0].value
        iterate += 1
    datasheet['B12']= sumvar/iterate

    datasheet['A13'] = "StandAbwPetCO2" # STDEV.S(limiter_beg2, limiter_end2), Column R
    r2_arr = []
    for l in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        r2_arr.append(l[0].value)
    datasheet['B13'] = statistics.stdev(r2_arr)

    datasheet['A14'] = "KorVE/SpO2" # Korrelation(limiter_beg1 (K), limiter_end1(K), limiter_beg2(P), limiter_end2(P)), Column K, P
    corr2=np.corrcoef(k2_arr, p2_arr)[0][1]
    datasheet['B14'] = corr2

    datasheet['A15'] = "KorVE/SpO2 t-Wert"
    t_wert = float(corr2 * math.sqrt(54))/(math.sqrt(1-(corr2**2)))
    datasheet['B15'] = t_wert

    datasheet['A16'] = "KorVE/SpO2 p-Wert"
    #datasheet['B15'] = "=T.DIST.2T(ABS(B14),52)" #excel-syntax  /kann openpyxl nicht parsen /TODO: fix
    datasheet['B16'] = t.ppf((1-(abs(t_wert)/2)), df=52)
    #print(t_wert)
    #print(t.ppf((abs(t_wert)), 52))
    #stats.t.ppf(1-alfa, free_deg)

    datasheet['A17'] = "RegVE/SPO2"
    datasheet['B17'] = np.polyfit(p2_arr, k2_arr,1)[0]
    return datasheet

def truedata_eval_hvr2(datasheet):
    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    if len(limiters) == 0:
        print("delimiter is zero in sheet" + str(globalcount))
        return datasheet
    if len(limiters) == 2:
        limiter_beg2 = limiters[0]
        limiter_end2 = limiters[1]
    #elif len(limiters) == 4:
    #    return data_eval_hvr1(datasheet)
    
    k2_arr = []
    p2_arr = []
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=11, max_col=11):
        k2_arr.append(val[0].value)
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=16, max_col=16):
        p2_arr.append(val[0].value)

    datasheet['A7'] = "MitHVRPetCO2" # Average(limiter_beg2, limiter_end2), Column R
    sumvar =0
    iterate=0
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        sumvar += val[0].value
        iterate += 1
    datasheet['B7']= sumvar/iterate

    datasheet['A8'] = "StandAbwPetCO2" # STDEV.S(limiter_beg2, limiter_end2), Column R
    r2_arr = []
    for l in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        r2_arr.append(l[0].value)
    datasheet['B8'] = statistics.stdev(r2_arr)
    
    datasheet['A9'] = "KorVE/SpO2" # Korrelation(limiter_beg1 (K), limiter_end1(K), limiter_beg2(P), limiter_end2(P)), Column K, P
    corr2=np.corrcoef(k2_arr, p2_arr)[0][1]
    datasheet['B9'] = corr2

    datasheet['A10'] = "KorVE/SpO2 t-Wert"
    t_wert = float(corr2 * math.sqrt(56))/(math.sqrt(1-(corr2**2)))             # warum auch immer hier 56 sein soll (ist wie in Excel)
    datasheet['B10'] = t_wert

    datasheet['A11'] = "KorVE/SpO2 p-Wert"
    #datasheet['B11'] = "=T.DIST.2T(ABS(B14),52)" #/TODO: check if correct

    datasheet['A12'] = "RegVE/SPO2"
    datasheet['B12'] = np.polyfit(p2_arr, k2_arr,1)[0]
    return datasheet
    
def truedata_eval_hvr3(datasheet):
    if (datasheet['D10'].value==None):
        return datasheet
    else:
        truedata_eval_hvr2(datasheet)

def truedata_eval_hcvr(datasheet):
    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    if len(limiters) != 2:
        print(globalcount)
        raise Exception("incorrect number of limiters HCVR")
    limiter_beg2 = limiters[0]
    limiter_end2 = limiters[1]
    k2_arr = []
    r2_arr = []
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=11, max_col=11):
        k2_arr.append(val[0].value)
    for val in datasheet.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        r2_arr.append(val[0].value)
        
    datasheet['A7'] = "KorPetCO2/VE"
    corr2=np.corrcoef(k2_arr, r2_arr)[0][1]
    datasheet['B7'] = corr2

    datasheet['A8'] = "KorVE/SpO2 t-Wert"
    t_wert = float(corr2 * math.sqrt(164))/(math.sqrt(1-(corr2**2)))
    datasheet['B8'] = t_wert

    datasheet['A9'] = "KorVE/SpO2 p-Wert"
    #datasheet['B9'] = "=T.DIST.2T(ABS(B14),52)" #/TODO: check if correct

    datasheet['A10'] = "RegPetCO2/VE"
    datasheet['B10'] = np.polyfit(r2_arr, k2_arr,1)[0]
    return datasheet

def data_eval_hvr1(datasheet):
    #print("hello from data_eval_hvr1")
    truedata = truedata_eval_hvr1(datasheet)
    rowtemp = 19
    for var in truedata.iter_rows(min_col=2, max_col=2, min_row=6, max_row=16):
        datasheet['B' + str(rowtemp)] = var[0].value
        rowtemp +=1

    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    if len(limiters) != 4:
        print(globalcount)
        raise Exception("incorrect number of limiters HVR1")
    limiter_beg1 = limiters[0]
    limiter_beg2 = limiters[2]
    limiter_end1 = limiters[1]
    limiter_end2 = limiters[3]

    color_field(datasheet, "cccccc", limiter_beg1, 11, limiter_end1, 18)
    color_field(datasheet, "b4c7dc", limiter_beg2 ,11, limiter_end2, 16)
    color_field(datasheet, "808080", limiter_beg2 ,18, limiter_end2, 18)

    datasheet['A6'] = "MitRuheVE"
    datasheet['B6'] = "=AVERAGE(K" + str(limiter_beg1) + ":K" + str(limiter_end1) + ")"

    datasheet['A7'] = "MitRuheBf"
    datasheet['B7'] = "=AVERAGE(L" + str(limiter_beg1) + ":L" + str(limiter_end1) + ")"

    datasheet['A8'] = "MitRuheVtL"
    datasheet['B8'] = "=AVERAGE(M" + str(limiter_beg1) + ":M" + str(limiter_end1) + ")"

    datasheet['A9'] = "MitRuheSpO2"
    datasheet['B9'] = "=AVERAGE(P" + str(limiter_beg1) + ":P" + str(limiter_end1) + ")"

    datasheet['A10'] = "MitRuhePetO2"
    datasheet['B10'] = "=AVERAGE(Q" + str(limiter_beg1) + ":Q" + str(limiter_end1) + ")"

    datasheet['A11'] = "MitRuhePetCO2"
    datasheet['B11'] = "=AVERAGE(R" + str(limiter_beg1) + ":R" + str(limiter_end1) + ")"

    datasheet['A12'] = "MitHVRPetCO2"
    datasheet['B12'] = "=AVERAGE(R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"

    datasheet['A13'] = "StandAbwPetCO2"
    datasheet['B13'] = "=STDEV(R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"

    datasheet['A14'] = "KorVE/SpO2"
    datasheet['B14'] = "=CORREL(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",P" + str(limiter_beg2) + ":P" + str(limiter_end2) + ")"

    datasheet['A15'] = "KorVE/SpO2 t-Wert"
    datasheet['B15'] = "=(B13*SQRT(54))/(SQRT(1-B13^2))"

    #BUG: t.sidt.2t is lower case, value can't be calculated; /TODO: fix
    datasheet['A16'] = "KorVE/SpO2 p-Wert"
    #datasheet['B15'] = "=T.DIST.2T(ABS(B14),52)"   #excel-syntax
    datasheet['B16'] = "=T.DIST(ABS(B14),52)"

    datasheet['A17'] = "RegVE/SPO2"
    datasheet['B17'] = "=LINEST(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",P" + str(limiter_beg2) + ":P" + str(limiter_end2) + ")"

    return datasheet

    
def data_eval_hvr2(datasheet):
    #print("hello from data_eval_hvr2")
    truedata = truedata_eval_hvr2(datasheet)
    rowtemp = 19
    for var in truedata.iter_rows(min_col=2, max_col=2, min_row=7, max_row=12):
        datasheet['B' + str(rowtemp)] = var[0].value
        rowtemp +=1
    
    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
        
    if len(limiters) == 2:
        limiter_beg2 = limiters[0]
        limiter_end2 = limiters[1]
    #elif len(limiters) == 4:
    #    return data_eval_hvr1(datasheet)
    else:
        raise Exception("incorrect number of limiters HVR2or3")

    color_field(datasheet, "b4c7dc", limiter_beg2 ,11, limiter_end2, 18)
    datasheet['A7'] = "MitHVRPetCO2"
    datasheet['B7'] = "=AVERAGE(R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"
    datasheet['A8'] = "StandAbwPetCO2"
    datasheet['B8'] = "=STDEV(R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"
    datasheet['A9'] = "KorVE/SpO2"
    datasheet['B9'] = "=CORREL(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",P" + str(limiter_beg2) + ":P" + str(limiter_end2) + ")"
    datasheet['A10'] = "KorVE/SpO2 t-Wert"
    datasheet['B10'] = "=(B9*SQRT(56))/(SQRT(1-B9^2))"
    datasheet['A11'] = "KorVE/SpO2 p-Wert"
    #datasheet['B11'] = "=T.DIST.2T(ABS(B10),54)"
    datasheet['A12'] = "RegVE/SPO2"
    datasheet['B12'] = "=LINEST(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",P" + str(limiter_beg2) + ":P" + str(limiter_end2) + ")"
    return datasheet

def data_eval_hvr3(datasheet):
    #print("hello from data_eval_hvr3")
    if (datasheet['D10'].value==None):
        return datasheet
    else:
        data_eval_hvr2(datasheet)
        truedata = truedata_eval_hvr2(datasheet)
        rowtemp = 19
        for var in truedata.iter_rows(min_col=2, max_col=2, min_row=7, max_row=12):
            datasheet['B' + str(rowtemp)] = var[0].value
            rowtemp +=1
    

def data_eval_hcvr(datasheet):
    #print("hello from data_eval_hcvr")
    print(globalcount)
    truedata = truedata_eval_hcvr(datasheet)
    rowtemp = 19
    for var in truedata.iter_rows(min_col=2, max_col=2, min_row=7, max_row=10):
        datasheet['B' + str(rowtemp)] = var[0].value
        rowtemp +=1
    
    iterate = 1
    limiters = []
    for row in datasheet.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    if len(limiters) != 2:
        print(globalcount)
        raise Exception("incorrect number of limiters HCVR")
    limiter_beg2 = limiters[0]
    limiter_end2 = limiters[1]

    color_field(datasheet, "b4c7dc", limiter_beg2 ,11, limiter_end2, 18)
    datasheet['A7'] = "KorPetCO2/VE"
    datasheet['B7'] = "=CORREL(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"
    datasheet['A8'] = "KorVE/SpO2 t-Wert"
    datasheet['B8'] = "=(B7*SQRT(164))/(SQRT(1-B7^2))"
    datasheet['A9'] = "KorVE/SpO2 p-Wert"
    #datasheet['B9'] = "=T.DIST.2T(ABS(B8),162)"
    datasheet['A10'] = "RegPetCO2/VE"
    datasheet['B10'] = "=LINEST(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"
    return datasheet


def data_eval_hcvr_with_mean_RuhePetCO2(datasheet_hcvr, datasheet_hvr1):
    meanPetCO2 = 0
    iterate = 1
    limiters = []
    for row in datasheet_hvr1.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    iterate = 0
    for row in datasheet_hvr1.iter_rows(min_col=18, max_col =18, min_row=limiters[0], max_row=limiters[1]):
        iterate += 1
        meanPetCO2 = meanPetCO2 + row[0].value
    meanPetCO2 = round(meanPetCO2/iterate)
    count = 0
    occ = []
    iterate = 1
    limiters = []
    for row in datasheet_hcvr.iter_rows(min_col=3, max_col=3):
        if (row[0].value == "x" or row[0].value == "X"):
            limiters.append(iterate)
        iterate +=1
    for row in datasheet_hcvr.iter_rows(min_col=18, max_col =18, max_row=limiters[1]):
        count += 1
        if row[0].value == meanPetCO2:
            occ.append(count)
    limiter_beg2 = occ[-1]
    limiter_end2 = limiters[1]
    k2_arr = []
    r2_arr = []
    for val in datasheet_hcvr.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=11, max_col=11):
        k2_arr.append(val[0].value)
    for val in datasheet_hcvr.iter_rows(min_row=limiter_beg2, max_row=limiter_end2, min_col=18, max_col=18):
        r2_arr.append(val[0].value)
    datasheet_hcvr['A27'] = "KorPetCO2/VE MeanRuhe_PetCO2"
    datasheet_hcvr['B27'] = "=CORREL(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"
    datasheet_hcvr['A28'] = "KorVE/SpO2 t-Wert MeanRuhe_PetCO2"
    datasheet_hcvr['B28'] = "=(B27*SQRT(164))/(SQRT(1-B27^2))"
    datasheet_hcvr['A29'] = "KorVE/SpO2 p-Wert MeanRuhe_PetCO2"
    #datasheet['B9'] = "=T.DIST.2T(ABS(B8),162)"
    datasheet_hcvr['A30'] = "RegPetCO2/VE MeanRuhe_PetCO2"
    datasheet_hcvr['B30'] = "=LINEST(K" + str(limiter_beg2) + ":K" + str(limiter_end2) + ",R" + str(limiter_beg2) + ":R" + str(limiter_end2) + ")"

    #truedata
    corr2 = np.corrcoef(k2_arr, r2_arr)[0][1]
    datasheet_hcvr['B46'] = corr2
    datasheet_hcvr['B47'] = float(corr2 * math.sqrt(164))/(math.sqrt(1-(corr2**2)))
    datasheet_hcvr['B48'] = ""
    datasheet_hcvr['B49'] = np.polyfit(r2_arr, k2_arr,1)[0]

    #/TODO: print to Resultfile

    return datasheet_hcvr

def create_resultlist (spreadfilelist): # TODO: test, if data fits ID
    filenames = get_filenames(datapath+"\Input")
    dict_toprinttoresultsheet = []
    dict_toprinttoresultsheet.append(['ID',['ID', filenames]])
    dict_toprinttoresultsheet.append(['HVR 1',[]])
    dict_toprinttoresultsheet.append(['HVR 2',[]])
    dict_toprinttoresultsheet.append(['HVR 3',[]])
    dict_toprinttoresultsheet.append(['HCVR',[]])
    i_hvr1 = 1
    i_hvr2 = 2
    i_hvr3 = 3
    i_hcvr = 4
    wbhvr3count = 0
    for wb in spreadfilelist:
        truedata_eval_hvr1(wb[hvr1])
        for row in wb[hvr1].iter_rows(min_row=4, max_row=30, min_col=1, max_col =2):
            if row[0].value != None:
                if row[0].value not in dict_toprinttoresultsheet[i_hvr1][1]:
                    dict_toprinttoresultsheet[i_hvr1][1].append(row[0].value)
                    dict_toprinttoresultsheet[i_hvr1][1].append([])
                dict_toprinttoresultsheet[i_hvr1][1][dict_toprinttoresultsheet[i_hvr1][1].index(row[0].value)+1].append(row[1].value)    

        truedata_eval_hvr2(wb[hvr2])
        for row in wb[hvr2].iter_rows(min_row=4, max_row=30, min_col=1, max_col =2):
            if row[0].value != None:
                if row[0].value not in dict_toprinttoresultsheet[i_hvr2][1]:
                    dict_toprinttoresultsheet[i_hvr2][1].append(row[0].value)
                    dict_toprinttoresultsheet[i_hvr2][1].append([])
                dict_toprinttoresultsheet[i_hvr2][1][dict_toprinttoresultsheet[i_hvr2][1].index(row[0].value)+1].append(row[1].value)    
        if hvr3 in wb.sheetnames and not wb[hvr3]['D10'].value==None:
            truedata_eval_hvr3(wb[hvr3])
            for row in wb[hvr3].iter_rows(min_row=4, max_row=30, min_col=1, max_col =2):
                if row[0].value != None:
                    if row[0].value not in dict_toprinttoresultsheet[i_hvr3][1]:
                        dict_toprinttoresultsheet[i_hvr3][1].append(row[0].value)
                        dict_toprinttoresultsheet[i_hvr3][1].append([])
                    for i in range(wbhvr3count):
                        dict_toprinttoresultsheet[i_hvr3][1][dict_toprinttoresultsheet[i_hvr3][1].index(row[0].value)+1].append(None)
                    dict_toprinttoresultsheet[i_hvr3][1][dict_toprinttoresultsheet[i_hvr3][1].index(row[0].value)+1].append(row[1].value)   
            wbhvr3count = 0 
        else:
            wbhvr3count += 1

        truedata_eval_hcvr(wb[hvcr])
        for row in wb[hvcr].iter_rows(min_row=4, max_row=30, min_col=1, max_col =2):
            if row[0].value != None:
                if row[0].value not in dict_toprinttoresultsheet[i_hcvr][1]:
                    dict_toprinttoresultsheet[i_hcvr][1].append(row[0].value)
                    dict_toprinttoresultsheet[i_hcvr][1].append([])
                dict_toprinttoresultsheet[i_hcvr][1][dict_toprinttoresultsheet[i_hcvr][1].index(row[0].value)+1].append(row[1].value)
    return dict_toprinttoresultsheet

def shift(inputchar):
    if inputchar[-1] == ',':
        return inputchar[:-1] + 'A'
    if inputchar[-1] == 'A':
        return inputchar[:-1] + 'B'
    if inputchar[-1] == 'B':
        return inputchar[:-1] + 'C'
    if inputchar[-1] == 'C':
        return inputchar[:-1] + 'D'
    if inputchar[-1] == 'D':
        return inputchar[:-1] + 'E'
    if inputchar[-1] == 'E':
        return inputchar[:-1] + 'F'
    if inputchar[-1] == 'F':
        return inputchar[:-1] + 'G'
    if inputchar[-1] == 'G':
        return inputchar[:-1] + 'H'
    if inputchar[-1] == 'H':
        return inputchar[:-1] + 'I'
    if inputchar[-1] == 'I':
        return inputchar[:-1] + 'J'
    if inputchar[-1] == 'J':
        return inputchar[:-1] + 'K'
    if inputchar[-1] == 'K':
        return inputchar[:-1] + 'L'
    if inputchar[-1] == 'L':
        return inputchar[:-1] + 'M'
    if inputchar[-1] == 'M':
        return inputchar[:-1] + 'N'
    if inputchar[-1] == 'N':
        return inputchar[:-1] + 'O'
    if inputchar[-1] == 'O':
        return inputchar[:-1] + 'P'
    if inputchar[-1] == 'P':
        return inputchar[:-1] + 'Q'
    if inputchar[-1] == 'Q':
        return inputchar[:-1] + 'R'
    if inputchar[-1] == 'R':
        return inputchar[:-1] + 'S'
    if inputchar[-1] == 'S':
        return inputchar[:-1] + 'T'
    if inputchar[-1] == 'T':
        return inputchar[:-1] + 'U'
    if inputchar[-1] == 'U':
        return inputchar[:-1] + 'V'
    if inputchar[-1] == 'V':
        return inputchar[:-1] + 'W'
    if inputchar[-1] == 'W':
        return inputchar[:-1] + 'X'
    if inputchar[-1] == 'X':
        return inputchar[:-1] + 'Y'
    if inputchar[-1] == 'Y':
        return inputchar[:-1] + 'Z'
    if inputchar[-1] == 'Z':
        if inputchar[0] == 'Z':
            strtemp = 'A'
            for i in range(0, len(inputchar)): strtemp += 'A'
            return strtemp
        inputchar = inputchar[:-1] + 'A'
        return shift(inputchar[:-1]) + inputchar[-1]
    return inputchar

def create_resulttable(workbooks):
    listtoparse=create_resultlist(workbooks)
    res = Workbook()
    rownumber = 1
    colnumber = "A"
    result_sheet = res.active
    for firstlevel in listtoparse:
        n1=2
        i1=1
        sheet_name=""
        for id_sheets in firstlevel:
            if isinstance((id_sheets), str) and id_sheets != "ID" and len(id_sheets) > 0 :
                sheet_name=id_sheets
            if i1%n1 ==0: 
                for val in id_sheets:
                    if isinstance(val, str):
                        rownumber = 1
                        if val != "ID":
                            result_sheet[colnumber + str(rownumber)].value = sheet_name + " - " + val
                        else:
                            result_sheet[colnumber + str(rownumber)].value = val
                    else:
                        for elem in val:
                            rownumber += 1
                            result_sheet[colnumber + str(rownumber)].value = elem
                        colnumber = shift(colnumber)
            else:
                i1 += 1

    res.save(datapath + "\Results.xlsx")
    return listtoparse
#--------------------------------
#-------function useage----------
#--------------------------------

list_of_spreadfiles = get_spreadsheets(datapath+"\Input")
for wb in list_of_spreadfiles:
    globalcount+=1
    if hvr1 in wb.sheetnames:
        data_eval_hvr1(wb[hvr1])
    if hvr2 in wb.sheetnames:
        data_eval_hvr2(wb[hvr2])
    if hvr3 in wb.sheetnames:
        data_eval_hvr3(wb[hvr3])

make_result_dirs(datapath, list_of_spreadfiles)
create_resulttable(list_of_spreadfiles)